# -*- coding: utf-8 -*-
"""
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import argrelextrema
import os
# import shutil
from scipy.optimize import curve_fit
from matplotlib.ticker import MultipleLocator
# from matplotlib import colors
# import matplotlib.font_manager as font_manager
import openpyxl
# from openpyxl.styles import Font, PatternFill
# import pickle
#%%

datapath='./'


for p_index in os.scandir(datapath):    
    if p_index.is_dir() and len(p_index.name) == 4:
        
        col_index = 1
        wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
        s2 = wb.active
        s2.title = 'periods'
        s2.cell(1,1).value = 'length (micron)'
        for i in range(16):
            s2.cell(i+2,1).value = round((i+8)*0.2,1)

        s3 = wb.create_sheet("lambda_N")  
        s3.cell(1,1).value = 'length (micron)'
        for i in range(16):
            s3.cell(i+2,1).value = round((i+8)*0.2,1)

        s3.cell(18,1).value = 'starting length'
        s3.cell(19,1).value = 'last length'
        s3.cell(20,1).value = '1st slope'
        s3.cell(21,1).value = '1st intercept'

        s3.cell(22,1).value = 'starting length'
        s3.cell(23,1).value = 'last length'
        s3.cell(24,1).value = '2nd slope'
        s3.cell(25,1).value = '2nd intercept'

        s4 = wb.create_sheet("I_Ratio")  
        s4.cell(1,1).value = 'length (micron)'
        for i in range(16):
            s4.cell(i+2,1).value = round((i+8)*0.2,1)




        fig = plt.figure(figsize=(10,3.75))

        ax1 = plt.subplot(131)#,)
        for axis in ['top','bottom','left','right']:
            ax1.spines[axis].set_linewidth(3.0)
        ax1.xaxis.set_minor_locator(MultipleLocator(.5))
        ax1.tick_params(which='minor', width=3.0, length=4.0)
        ax1.tick_params(which='major', width=3.0, length=6.0)
        ax1.xaxis.set_tick_params(width=3)
        ax1.yaxis.set_tick_params(width=3)
        ax1.set_xlabel('Length ($\mu$m)',rotation=0,fontname="Arial", fontsize=24,color='w')
        ax1.set_xticks([2.0,3.0,4.0,5.0])
        ax1.set_xlim([1.5,5.0])
        ax1.set_ylim([0,30])
        ax1.set_yticks([0,10,20,30])
        ax1.yaxis.set_minor_locator(MultipleLocator(5))
        # ax1.set_aspect(1.5*3.4/51)
        plt.xticks(fontname="Arial", fontsize=24)
        plt.yticks(fontname="Arial", fontsize=24)
        ax1.set_title('Periods (s)',fontname="Arial", fontsize=24)

        ax1.yaxis.set_label_coords(-0.5, .77)

        ax2 = plt.subplot(132)
        for axis in ['top','bottom','left','right']:
            ax2.spines[axis].set_linewidth(3.0)
        ax2.xaxis.set_minor_locator(MultipleLocator(.5))
        ax2.tick_params(which='minor', width=3.0, length=4.0)
        ax2.tick_params(which='major', width=3.0, length=6.0)
        ax2.xaxis.set_tick_params(width=3)
        ax2.yaxis.set_tick_params(width=3)
        ax2.set_xlabel('Length ($\mu$m)',rotation=0,fontname="Arial", fontsize=24,color='w')
        ax2.set_xticks([2.0,3.0,4.0,5.0])
        ax2.set_xlim([1.5,5.0])
        ax2.set_ylim([0.0,4.0])
        ax2.set_yticks([0,1,2,3,4])
        ax2.yaxis.set_minor_locator(MultipleLocator(0.5))
        # ax2.set_aspect(1.5*3.4/3.7)
        plt.xticks(fontname="Arial", fontsize=24)
        plt.yticks(fontname="Arial", fontsize=24)
        ax2.set_title('$\lambda_N$',rotation=0,fontname="Arial", fontsize=24)

        ax3 = plt.subplot(133)
        for axis in ['top','bottom','left','right']:
            ax3.spines[axis].set_linewidth(3.0)
        ax3.xaxis.set_minor_locator(MultipleLocator(.5))
        ax3.tick_params(which='minor', width=3.0, length=4.0)
        ax3.tick_params(which='major', width=3.0, length=6.0)
        ax3.xaxis.set_tick_params(width=3)
        ax3.yaxis.set_tick_params(width=3)
        ax3.set_xlabel('Length ($\mu$m)',rotation=0,fontname="Arial", fontsize=24,color='w')
        ax3.set_xticks([2.0,3.0,4.0,5.0])
        ax3.set_xlim([1.5,5.0])
        ax3.set_ylim([0.2,0.8])
        ax3.set_yticks([0.2,0.4,0.6,0.8])
        ax3.yaxis.set_minor_locator(MultipleLocator(0.1))

        # ax3.set_aspect(1.5*3.4/0.6)
        plt.xticks(fontname="Arial", fontsize=24)
        plt.yticks(fontname="Arial", fontsize=24)
        ax3.set_title('$I_{Ratio}$',rotation=0,fontname="Arial", fontsize=24)

        # ax3.set_yticks([0,0.2,0.4,0.6,0.8,1.0])
        # ax3.set_xticks([2,3,4])
        # ax3.set_ylim([0.3,0.85])
        # plt.xticks(fontname="Arial", fontsize=24)
        # plt.yticks(fontname="Arial", fontsize=24)
        # plt.ylabel('$I_{Ratio}$',rotation=90,fontname="Arial", fontsize=24)

        # ax2.tick_params(axis='y',colors='red')

        # ax.tick_params(axis='y',colors='blue')

        # ax.spines["left"].set_edgecolor("blue")
        # ax3.spines["left"].set_linewidth(3.0)

        # ax2.spines["right"].set_edgecolor("red")
        # ax3.spines["right"].set_linewidth(3.0)

        
        col_index += 1
        
        cell_length = []
        periods = []
        lambda0 = []
        Iratio = []
        lambda_D = p_index.name
        
        a0all = []
        b0all = []
        c0all = []
        
        for numberx in range(8,24):
            Nx=numberx
            # print(Nx)
            cell_length.append(Nx*0.2)    
            filehead=datapath+"/"+p_index.name+"/"+str(numberx).zfill(2)+"/"
            tspan=np.load(filehead+"time.npy")
            xspan=np.load(filehead+"x.npy")
            DD=np.load(filehead+"DD.npy")
            DT=np.load(filehead+"DT.npy")
            E=np.load(filehead+"E.npy")
            d=np.load(filehead+"d.npy")
            de=np.load(filehead+"de.npy")
            
            #%% find the peak of all time, all position after 40 seconds
            start_index = np.where(tspan > 40)[0][0]
            
            dmem = d + de
            
            all_position_diff_time_max = np.max(dmem[:,start_index:],axis=1)
            
            all_position_diff_time_max_index = np.argmax(dmem[:,start_index:],axis=1)
            
            max_position_index = np.argmax(all_position_diff_time_max)
            max_time_index = all_position_diff_time_max_index[np.argmax(all_position_diff_time_max)]

            
            #%% find snapshot with max to get I ratio
            
            upper_position_diff_time_max = np.max(dmem[:int(Nx/2),start_index:],axis=1)
            upper_position_diff_time_max_index = np.argmax(dmem[:int(Nx/2),start_index:],axis=1)
            upper_max_position_index = np.argmax(upper_position_diff_time_max)
            upper_max_time_index = upper_position_diff_time_max_index[upper_max_position_index]
            
            lower_position_diff_time_max = np.max(dmem[int(Nx/2):,start_index:],axis=1)
            lower_position_diff_time_max_index = np.argmax(dmem[int(Nx/2):,start_index:],axis=1)
            lower_max_position_index = np.argmax(lower_position_diff_time_max)
            lower_max_time_index = lower_position_diff_time_max_index[lower_max_position_index]
                        
            upper_curve = dmem[:,upper_max_time_index+start_index]
            lower_curve = dmem[:,lower_max_time_index+start_index]
            
            max_int = np.max([upper_curve,lower_curve])
            
            left_curve = upper_curve/max_int
            
            right_curve = lower_curve/max_int
            
            solid_curve = np.max(np.vstack((left_curve,right_curve)),axis=0)
            
            if Nx % 2 == 0:
                mid_region_index = [int(Nx/2)-1,int(Nx/2)]
            else:
                mid_region_index = [int(Nx/2)]
            
            Iratio.append(np.min(solid_curve[mid_region_index]))
                
            
            #%% get periods
            time_curve = dmem[max_position_index,start_index:]
            
            peak_index = argrelextrema(time_curve, np.greater)[0] + start_index
            
            getperiod = False
            
            if len(peak_index) < 2:
                periods.append(0)
                getperiod = True

            elif len(peak_index) ==2 : # 只看倒數第二個 peak之後的 dynamic range，可以幫忙排除 damping case
                dynamic_curve = dmem[max_position_index,peak_index[-2]:]
                dynamic_range = 100 * (np.max(dynamic_curve) - np.min(dynamic_curve)) / np.mean(dynamic_curve)
                if dynamic_range < 5: # 5 % 可以再提高看看，可是目前沒有看到 cases
                    periods.append(0)
                    getperiod = True
            
            else : # 只看倒數第三個 peak之後的 dynamic range，可以幫忙排除 damping case
                dynamic_curve = dmem[max_position_index,peak_index[-3]:]
                dynamic_range = 100 * (np.max(dynamic_curve) - np.min(dynamic_curve)) / np.mean(dynamic_curve)
                if dynamic_range < 5: # 5 % 可以再提高看看，可是目前沒有看到 cases
                    periods.append(0)
                    getperiod = True

            if not getperiod:
                # remove local peak by peak_value_cv, 3.7 %
                peak_index = list(peak_index)
                peak_mean_value = np.mean(dmem[max_position_index,peak_index])
                peak_std_value = np.std(dmem[max_position_index,peak_index])
                peak_cv = 100 * peak_std_value / peak_mean_value
                while peak_cv > 3.5 and len(peak_index) > 2:
                    peak_index.pop(np.argmin(dmem[max_position_index,peak_index]))
                    peak_mean_value = np.mean(dmem[max_position_index,peak_index])
                    peak_std_value = np.std(dmem[max_position_index,peak_index])
                    peak_cv = 100 * peak_std_value / peak_mean_value
                    # plt.figure()
                    # plt.title(p_index.name+' '+ str(Nx) +' '+ str(round(peak_cv,4)))
                    # plt.plot(tspan[start_index:],time_curve)
                    # plt.plot(tspan[peak_index],dmem[max_position_index,peak_index],'o')
                
                # unify peaks
                currents = []
                for i in range(len(peak_index)-1):
                    currents.append(np.abs(tspan[peak_index[i+1]] - tspan[peak_index[i]]))
                current_mean = np.mean(currents)
                current_std  = np.std(currents)
                current_cv = 100 * current_std / current_mean
                
                # if current_cv > 2:
                #     plt.figure()
                #     plt.title(p_index.name+' '+ str(Nx) + ' '+ str(round(current_cv,4)))
                #     plt.plot(tspan[start_index:],time_curve)
                #     plt.plot(tspan[peak_index],dmem[max_position_index,peak_index],'o')
                    
                old_cv = current_cv
                old_peaks = peak_index.copy()
                while len(currents) > 2 and current_cv <= old_cv :
                    old_cv = current_cv
                    old_peaks = peak_index.copy()
                    
                    pop_index = np.argmin(currents)
                    if pop_index == len(currents)-1:
                        peak_index.pop(pop_index+1)  
                    elif currents[pop_index+1] > currents[pop_index-1]:  
                        peak_index.pop(pop_index)
                    else:
                        peak_index.pop(pop_index+1)
                        
                    currents = []
                    for i in range(len(peak_index)-1):
                        currents.append(np.abs(tspan[peak_index[i+1]] - tspan[peak_index[i]]))
                    current_mean = np.mean(currents)
                    current_std  = np.std(currents)
                    current_cv = 100 * current_std / current_mean
                    
                    # plt.figure()
                    # plt.title(p_index.name+' '+ str(Nx) + ' '+ str(round(current_cv,4)))
                    # plt.plot(tspan[start_index:],time_curve)
                    # plt.plot(tspan[peak_index],dmem[max_position_index,peak_index],'o')
                    
                if old_cv < current_cv:
                    peak_index = old_peaks.copy()
                    current_cv = old_cv
                # plt.figure()
                # plt.title(p_index.name+' '+ str(Nx) + ' '+ str(round(old_cv,4)))
                # plt.plot(tspan[start_index:],time_curve)
                # plt.plot(tspan[peak_index],dmem[max_position_index,peak_index],'o')
                
                currents = []
                if len(peak_index) == 1:
                    currents.append(0)
                else:
                    for i in range(len(peak_index)-1):
                        currents.append(np.abs(tspan[peak_index[i+1]] - tspan[peak_index[i]]))

                periods.append(np.mean(currents))


            #%% get lambda
            # print(len(cell_length),len(periods))
            
            # all_position_diff_time_max = np.max(dmem[:,start_index:],axis=1)
            
            # all_position_diff_time_max_index = np.argmax(dmem[:,start_index:],axis=1)
            
            # max_position_index = np.argmax(all_position_diff_time_max)
            # max_time_index = all_position_diff_time_max_index[max_position_index]

            curve = dmem[:,max_time_index+start_index]
            # posi = np.linspace(0.0,1.0,len(curve))
            posi = np.linspace(0.1/(Nx*0.2),(Nx*0.2-0.1)/(Nx*0.2),len(curve))
            
            # turn to from upper-left to down-right
            if max_position_index > numberx*0.5:
                curve = curve[::-1]
            
            #%% fitting lambda
            if np.argmax(curve) < np.argmin(curve):
                y = curve[np.argmax(curve):np.argmin(curve)+1]
            else:
                y = curve[np.argmin(curve):np.argmax(curve)+1]
            curve_max = np.max(y)
            y = y / curve_max
            # print('Nx: ',Nx,len(y))
            x0 = []
            for x0_index in range(len(y)):
                x0.append(posi[x0_index])
            
            if np.min(y)/np.max(y) > 0.95:
                lambda0.append(0)
                a0all.append(1)
                b0all.append(0)
                c0all.append(0)
            else:
                try:
                    popt0, pcov0 = curve_fit(lambda t, a, b, c: a * np.exp(-b * t) + c, x0, y)
                    a0 = popt0[0]
                    b0 = popt0[1]
                    c0 = popt0[2]
                except:
                    a0 = 1
                    b0 = 0
                    c0 = 0

                lambda0.append(b0)
                a0all.append(a0)
                b0all.append(b0)
                c0all.append(c0)

        lambda0 = np.array(np.abs(lambda0)) # 有時拿到負的，都轉成正的
        all_length = np.array(cell_length)
        periods = np.array(periods)
        
        Iratio = []
        for Lindex in range(16):
            imax = a0all[Lindex] + c0all[Lindex]
            imin = a0all[Lindex] * np.exp(-b0all[Lindex]*0.5) + c0all[Lindex]
            Iratio.append(imin/imax)
        
        s2.cell(1,col_index).value = 'period (sec)'
        # s2.cell(1,col_index).fill = PatternFill(fill_type="solid", fgColor="DDDD00") 
        s3.cell(1,col_index).value = 'lambda_N' #float(p_index.name)
        # s3.cell(1,col_index).fill = PatternFill(fill_type="solid", fgColor="DDDD00") 
        s4.cell(1,col_index).value = 'I_Ratio' #float(p_index.name)
        # s4.cell(1,col_index).fill = PatternFill(fill_type="solid", fgColor="DDDD00") 
        
        for i in range(16):
            s2.cell(i+2,col_index).value = round(periods[i],3)
        
        for i in range(16):
            if periods[i] > 0:
                s3.cell(i+2,col_index).value = round(lambda0[i],3)
            else:
                s3.cell(i+2,col_index).value = 0

        for i in range(16):
            if periods[i] > 0:
                s4.cell(i+2,col_index).value = round(Iratio[i],3)
            else:
                s4.cell(i+2,col_index).value = 1
        
        to_fit1 = periods > 0
        to_fit2 = lambda0 > 0
        to_fit = to_fit1 & to_fit2
        lambda0 = np.array(lambda0)[to_fit]
        cell_length = np.array(cell_length)[to_fit]
        Iratio = np.array(Iratio)[to_fit]
        
        line1, = ax1.plot(all_length,periods,'-o',label = p_index.name,markersize=6,linewidth=2,
                           zorder=10-col_index,clip_on=False,
                          )
        
        line2 = ax2.plot(cell_length,lambda0,'o',markersize=6,
                         zorder=10-col_index,clip_on=False,
                         )

        model2 = np.poly1d(np.polyfit(cell_length, lambda0, 2))
        errsum = []
        for gindex in range(3,len(cell_length)-2):
            # print(len(cell_length),len(cell_length[:gindex]),len(cell_length[gindex:]))
            modelshort = np.poly1d(np.polyfit(cell_length[:gindex], lambda0[:gindex], 1))
            modellong = np.poly1d(np.polyfit(cell_length[gindex:], lambda0[gindex:], 1))
            errsum.append(np.sum(np.abs(modelshort(cell_length[:gindex])-lambda0[:gindex]))+
                  np.sum(np.abs( modellong(cell_length[gindex:])-lambda0[gindex:])))
        # print(errsum,np.argmin(errsum))
        gtarget = np.argmin(errsum)+3
        modelshort = np.poly1d(np.polyfit(cell_length[:gtarget], lambda0[:gtarget], 1))
        modellong = np.poly1d(np.polyfit(cell_length[gtarget:], lambda0[gtarget:], 1))
        # print(np.sum(np.abs(modelshort(cell_length[:gtarget])-lambda0[:gtarget]))+
        #       np.sum(np.abs( modellong(cell_length[gtarget:])-lambda0[gtarget:])))
        
        s3.cell(18,col_index).value = cell_length[0]
        s3.cell(19,col_index).value = cell_length[gtarget-1]
        s3.cell(20,col_index).value = modelshort.coefficients[0]
        s3.cell(21,col_index).value = modelshort.coefficients[1]
        s3.cell(22,col_index).value = cell_length[gtarget]
        s3.cell(23,col_index).value = cell_length[-1]
        s3.cell(24,col_index).value = modellong.coefficients[0]
        s3.cell(25,col_index).value = modellong.coefficients[1]
        
        
        
        polyline = np.linspace(np.min(cell_length[:gtarget]), np.max(cell_length[:gtarget])+0.05, 50)
        ax2.plot(polyline, modelshort(polyline), color=line2[0].get_color(),linewidth=2)
        
        polyline = np.linspace(np.min(cell_length[gtarget:])-0.05, np.max(cell_length[gtarget:]), 50)
        ax2.plot(polyline, modellong(polyline), color=line2[0].get_color(),linewidth=2)
        
        
        line3=ax3.plot(cell_length, Iratio,'o',markersize=6,
                       zorder=10-col_index,clip_on=False)
                           
        model2 = np.poly1d(np.polyfit(cell_length, Iratio, 2))
        polyline = np.linspace(np.min(cell_length), np.max(cell_length), 50)
        ax3.plot(polyline, model2(polyline),'-', color=line3[0].get_color(),linewidth=2)
                
        print(p_index.name)#,round(np.max(lambda0),2),round(np.max(Iratio),2))

#%%
        fig.tight_layout(pad=0.5)
        plt.savefig('ana.'+p_index.name+'.png', dpi=140)
    
        wb.save(p_index.name +'.xlsx')     
